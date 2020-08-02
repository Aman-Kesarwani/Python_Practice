# D:\Workspace\Python_Practice\transactions\transactions.xlsx

import openpyxl

#wb = penpyxl.Workbook()

wb = openpyxl.load_workbook(
    r"D:\Workspace\Python_Practice\transactions\transactions.xlsx")

print(wb.sheetnames)

sheet = wb["Sheet1"]

cell = sheet["a1"]

print(cell.row)
print(cell.column)
print(cell.coordinate)

print(sheet.max_row)
print(sheet.max_column)

for row in range(1, sheet.max_row+1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)

column = sheet["a"]
print("Column Values", column)

cells = sheet["a:c"]
print("a to c", cells)
cells_1 = sheet["a1:c3"]
print("cells_1", cells_1)

print("row", sheet[1:3])

sheet.append([101, 102, 103])
wb.save("TRANSACTIONS.CSV")
